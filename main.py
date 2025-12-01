import os
from typing import Dict, List, TypedDict, Any
from langgraph.graph import StateGraph, END
from langchain_ollama import ChatOllama
from langchain_core.prompts import PromptTemplate
from langchain_core.output_parsers import JsonOutputParser
from pydantic import BaseModel, Field
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import json
import re

print("✅ Loading libraries...")


# ---------------------------------------------------------------------------
# 1) STATE DEFINITION
# ---------------------------------------------------------------------------

class TaskDetails(BaseModel):
    name: str = Field(description="Task name or brief description.")
    start_time: str = Field(description="Task start time (e.g., 10:00 AM).")
    duration: str = Field(description="Task duration (e.g., 1 hour or 30 minutes).")
    schedule_date: str = Field(description="Scheduled date or day for the task (e.g., today, tomorrow, 2025-12-01).")

class ScheduleState(TypedDict):
    user_input: str
    tasks: List[Dict[str, Any]]
    current_task_index: int
    completed_tasks: List[Dict[str, Any]]
    schedule_date: str
    conflict_detected: bool
    is_available: bool
    llm_response: str
    new_user_input: str
    extraction_failed: bool
    excel_file_path: str
    conflict_details: Dict[str, Any]
    max_retries: int
    current_retry: int
    waiting_for_user_input: bool

# ---------------------------------------------------------------------------
# 2) LLM SETUP
# ---------------------------------------------------------------------------

try:
    llm = ChatOllama(model="llama3", temperature=0.0)
    parser = JsonOutputParser(pydantic_object=TaskDetails)
except Exception as e:
    print(f"⚠️ Warning: {e}")
    llm = None
    parser = None

# ---------------------------------------------------------------------------
# 3) EXCEL MANAGEMENT
# ---------------------------------------------------------------------------

def initialize_excel_file(file_path: str = "schedule_tasks.xlsx"):
    """Initialize Excel file if it doesn't exist"""
    try:
        if not os.path.exists(file_path):
            df = pd.DataFrame(columns=[
                "Task Name", 
                "Start Time", 
                "Duration", 
                "Date", 
                "Status",
                "Added Time"
            ])
            df.to_excel(file_path, index=False, engine='openpyxl')
            print(f"✅ Created new Excel file: {file_path}")
        return file_path
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        return file_path

def save_task_to_excel(task: Dict[str, Any], schedule_date: str, file_path: str):
    """Save task to Excel file"""
    try:
        if not os.path.exists(file_path):
            initialize_excel_file(file_path)
        
        df = pd.read_excel(file_path, engine='openpyxl')
        
        new_task = {
            "Task Name": task.get("name", "Not specified"),
            "Start Time": task.get("start_time", "Not specified"),
            "Duration": task.get("duration", "Not specified"),
            "Date": schedule_date,
            "Status": task.get("status", "completed"),
            "Added Time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        df = pd.concat([df, pd.DataFrame([new_task])], ignore_index=True)
        df.to_excel(file_path, index=False, engine='openpyxl')
        
        apply_excel_formatting(file_path)
        
        print(f"✅ Task saved to Excel")
        return True
        
    except Exception as e:
        print(f"❌ Error saving task to Excel: {e}")
        return False

def apply_excel_formatting(file_path: str):
    """Apply formatting to Excel file"""
    try:
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        column_widths = {'A': 30, 'B': 15, 'C': 15, 'D': 15, 'E': 12, 'F': 20}
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        workbook.save(file_path)
        
    except Exception as e:
        print(f"⚠️ Could not apply formatting: {e}")

def read_all_tasks_from_excel(file_path: str):
    """Read all tasks from Excel file"""
    try:
        if os.path.exists(file_path):
            return pd.read_excel(file_path, engine='openpyxl')
        return pd.DataFrame()
    except Exception as e:
        print(f"❌ Error reading Excel: {e}")
        return pd.DataFrame()

def get_existing_tasks(file_path: str, date: str, time: str):
    """Get existing tasks at the same date and time"""
    try:
        df = read_all_tasks_from_excel(file_path)
        if df.empty:
            return []
        
        conflicting_tasks = []
        for _, task in df.iterrows():
            if (str(task['Date']) == str(date) and 
                str(task['Start Time']) == str(time)):
                conflicting_tasks.append({
                    'name': task['Task Name'],
                    'time': task['Start Time'],
                    'date': task['Date']
                })
        
        return conflicting_tasks
    except Exception as e:
        print(f"❌ Error searching for tasks: {e}")
        return []

# ---------------------------------------------------------------------------
# 4) IMPROVED EXTRACTION WITH FALLBACK
# ---------------------------------------------------------------------------

def manual_extract_tasks(user_input: str):
    """Manual task extraction if LLM fails"""
    try:
        # Patterns to find time and date
        time_patterns = [
            r'at\s*(\d+(?::\d+)?\s*(?:AM|PM|am|pm)?)',
            r'(\d+(?::\d+)?\s*(?:AM|PM|am|pm)?)\s*',
            r'time\s*(\d+(?::\d+)?)',
            r'(\d+(?::\d+)?)\s*(?:o\'clock|clock)',
            r'(\d+)\s*(?:AM|PM|am|pm)'
        ]
        
        date_patterns = [
            r'(\d{1,2}-\d{1,2}-\d{4})',
            r'(\d{1,2}/\d{1,2}/\d{4})',
            r'(today|tomorrow|day after tomorrow|next week|monday|tuesday|wednesday|thursday|friday|saturday|sunday)',
            r'on\s*(\w+\s+\d{1,2})'
        ]
        
        # Extract time
        start_time = "Not specified"
        for pattern in time_patterns:
            match = re.search(pattern, user_input, re.IGNORECASE)
            if match:
                start_time = match.group(1).strip()
                # Add AM/PM if missing
                if not re.search(r'(AM|PM|am|pm)', start_time, re.IGNORECASE):
                    if 'am' in user_input.lower() or 'morning' in user_input.lower():
                        start_time += ' AM'
                    elif 'pm' in user_input.lower() or 'afternoon' in user_input.lower() or 'evening' in user_input.lower():
                        start_time += ' PM'
                break
        
        # Extract date
        schedule_date = "today"
        for pattern in date_patterns:
            match = re.search(pattern, user_input, re.IGNORECASE)
            if match:
                schedule_date = match.group(1).strip()
                # Handle "day after tomorrow"
                if 'day after tomorrow' in user_input.lower():
                    schedule_date = "day after tomorrow"
                break
        
        # Extract task name (remove time and date patterns)
        name = user_input
        for pattern in time_patterns + date_patterns:
            name = re.sub(pattern, '', name, flags=re.IGNORECASE)
        
        # Remove common scheduling words
        name = re.sub(r'\b(at|on|time|appointment|meeting|schedule)\b', '', name, flags=re.IGNORECASE)
        name = re.sub(r'\s+', ' ', name).strip()
        
        if not name or name == 'my with is':
            name = "Dentist Appointment" if 'dentist' in user_input.lower() else "New task"
        elif 'dentist' in user_input.lower() and 'appointment' not in name.lower():
            name = "Dentist Appointment"
        
        return {
            "name": name,
            "start_time": start_time,
            "duration": "1 hour",
            "schedule_date": schedule_date
        }
    except Exception as e:
        print(f"❌ Error in manual extraction: {e}")
        return {
            "name": "New task",
            "start_time": "Not specified",
            "duration": "1 hour",
            "schedule_date": "today"
        }

def extract_variables(state: ScheduleState) -> ScheduleState:
    """Extract task details from input"""
    print("🔍 Extracting task details...")
    
    input_to_parse = state.get("new_user_input") or state["user_input"]
    
    if input_to_parse == "FAILED_EXTRACTION":
        input_to_parse = state["user_input"]
    
    # Use manual extraction if LLM is not available or if we're in retry mode
    if llm is None or state.get("waiting_for_user_input", False):
        print("⚠️ Using manual extraction")
        extracted_task = manual_extract_tasks(input_to_parse)
        
        tasks = state.get("tasks", [])
        if state["current_task_index"] < len(tasks):
            tasks[state["current_task_index"]] = extracted_task
        else:
            tasks.append(extracted_task)
            
        return {
            "tasks": tasks,
            "schedule_date": extracted_task.get("schedule_date", "today"),
            "new_user_input": "",
            "extraction_failed": False,
            "waiting_for_user_input": False
        }
    
    # Use LLM if available
    prompt = PromptTemplate(
        template="""You are an expert task scheduling assistant. Extract task details from the user request.

Request: {request}

Output ONLY JSON in this format:
{{
  "name": "Task name",
  "start_time": "Time",
  "duration": "Duration", 
  "schedule_date": "Date"
}}

If no date is specified, use 'today'
If no duration is specified, use '1 hour'
Output ONLY JSON without any additional text""",
        input_variables=["request"],
        partial_variables={"format_instructions": parser.get_format_instructions() if parser else ""},
    )
    
    chain = prompt | llm | parser
    
    try:
        llm_output = chain.invoke({"request": input_to_parse})
        
        task = {
            "name": llm_output.get("name", "Unspecified task"),
            "start_time": llm_output.get("start_time", "Not specified"),
            "duration": llm_output.get("duration", "1 hour"),
            "status": "pending"
        }
        
        tasks = state.get("tasks", [])
        if state["current_task_index"] < len(tasks):
            tasks[state["current_task_index"]] = task
        else:
            tasks.append(task)
            
        return {
            "tasks": tasks,
            "schedule_date": llm_output.get("schedule_date", "today"),
            "new_user_input": "",
            "extraction_failed": False,
            "waiting_for_user_input": False
        }
        
    except Exception as e:
        print(f"❌ Error in automatic extraction: {e}")
        print("🔄 Using manual extraction...")
        
        # Use manual extraction as fallback
        extracted_task = manual_extract_tasks(input_to_parse)
        
        tasks = state.get("tasks", [])
        if state["current_task_index"] < len(tasks):
            tasks[state["current_task_index"]] = extracted_task
        else:
            tasks.append(extracted_task)
            
        return {
            "tasks": tasks,
            "schedule_date": extracted_task.get("schedule_date", "today"),
            "new_user_input": "",
            "extraction_failed": False,
            "waiting_for_user_input": False
        }

def check_conflict(state: ScheduleState) -> ScheduleState:
    """Check for conflicts with existing tasks"""
    print("⚡ Checking for conflicts...")
    
    if not state.get("tasks") or state["current_task_index"] >= len(state["tasks"]):
        return {"conflict_detected": False}
    
    current_task = state["tasks"][state["current_task_index"]]
    current_time = current_task["start_time"]
    current_date = state["schedule_date"]
    
    # Normalize time format for comparison
    normalized_time = current_time.replace(' ', '').upper()
    if 'AM' not in normalized_time and 'PM' not in normalized_time:
        normalized_time += 'AM'  # Default to AM
    
    # Search for conflicting tasks
    conflicting_tasks = get_existing_tasks(state['excel_file_path'], current_date, current_time)
    
    if conflicting_tasks:
        print(f"⚠️ Conflict detected with {len(conflicting_tasks)} tasks")
        return {
            "conflict_detected": True,
            "conflict_details": {
                "current_task": current_task,
                "conflicting_tasks": conflicting_tasks,
                "conflict_count": len(conflicting_tasks)
            }
        }
    
    return {"conflict_detected": False}

def check_availability(state: ScheduleState) -> ScheduleState:
    """Check availability"""
    print("🔎 Checking availability...")
    
    if not state.get("tasks") or state["current_task_index"] >= len(state["tasks"]):
        return {"is_available": True}
    
    current_task = state["tasks"][state["current_task_index"]]
    
    # Simulate unavailability for "team meeting"
    if "team meeting" in current_task["name"].lower():
        print("❌ Not available for 'team meeting'")
        return {"is_available": False}
        
    return {"is_available": True}

def output_schedule(state: ScheduleState) -> ScheduleState:
    """Save task and output result"""
    print("📅 Scheduling task...")
    
    if not state.get("tasks") or state["current_task_index"] >= len(state["tasks"]):
        return {
            "llm_response": "❌ No current task",
            "new_user_input": ""
        }
    
    current_task = state["tasks"][state["current_task_index"]]
    current_task["status"] = "completed"
    
    completed = state.get("completed_tasks", [])
    completed.append(current_task)
    
    # Save to Excel
    excel_saved = save_task_to_excel(current_task, state['schedule_date'], state['excel_file_path'])
    
    success_message = f"✅ Task scheduled: **{current_task['name']}** at **{current_task['start_time']}** on **{state['schedule_date']}** successfully."
    
    if excel_saved:
        success_message += f"\n💾 Saved to Excel: {state['excel_file_path']}"
    
    return {
        "completed_tasks": completed,
        "current_task_index": state["current_task_index"] + 1,
        "llm_response": success_message,
        "new_user_input": "",
        "current_retry": 0,  # Reset counter
        "waiting_for_user_input": False
    }

def ask_for_new_time(state: ScheduleState) -> ScheduleState:
    """Ask for new time with conflict details"""
    print("🔄 Requesting new time...")
    
    current_retry = state.get("current_retry", 0) + 1
    max_retries = state.get("max_retries", 3)
    
    if current_retry > max_retries:
        return {
            "llm_response": "🚫 Maximum attempts exceeded. Please try again later.",
            "new_user_input": "",
            "current_retry": 0,
            "waiting_for_user_input": False
        }
    
    if not state.get("tasks") or state["current_task_index"] >= len(state["tasks"]):
        return {
            "llm_response": "❌ No current task",
            "new_user_input": ""
        }
    
    current_task = state["tasks"][state["current_task_index"]]
    conflict_details = state.get("conflict_details", {})
    
    if state.get("conflict_detected"):
        conflicting_tasks = conflict_details.get("conflicting_tasks", [])
        
        if conflicting_tasks:
            conflict_message = f"\n🚨 **Time conflict detected!**\n"
            conflict_message += f"⏰ Requested time: {current_task['start_time']} on {state['schedule_date']}\n\n"
            conflict_message += "📋 **Conflicting tasks:**\n"
            
            for i, task in enumerate(conflicting_tasks, 1):
                conflict_message += f"   {i}. {task['name']} - {task['time']}\n"
            
            conflict_message += f"\n💡 **Please enter a new time or date (Attempt {current_retry}/{max_retries}):**"
            
            return {
                "llm_response": conflict_message,
                "conflict_detected": False,
                "is_available": True,
                "new_user_input": "",
                "current_retry": current_retry,
                "waiting_for_user_input": True  # Signal that we need user input
            }
    
    return {
        "llm_response": "❌ Could not schedule the task. Please try again.",
        "new_user_input": "",
        "current_retry": current_retry,
        "waiting_for_user_input": True
    }

# ---------------------------------------------------------------------------
# 5) CONDITIONAL LOGIC
# ---------------------------------------------------------------------------

def route_conflict_check(state: ScheduleState) -> str:
    """Route based on conflict check result"""
    if state.get("extraction_failed", False):
        return "ask_new_time"
    elif state["conflict_detected"]:
        return "ask_new_time"
    else:
        return "check_availability"

def route_availability_check(state: ScheduleState) -> str:
    """Route based on availability check result"""
    if state.get("is_available", True):
        return "output_schedule"
    else:
        return "ask_new_time"

def route_after_ask_new_time(state: ScheduleState) -> str:
    """Route after asking for new time"""
    if state.get("waiting_for_user_input", False):
        return "END"  # Stop the graph and wait for user input
    else:
        return "extract_variables"

# ---------------------------------------------------------------------------
# 6) BUILD WORKFLOW
# ---------------------------------------------------------------------------

def build_workflow():
    """Build workflow"""
    workflow = StateGraph(ScheduleState)

    workflow.add_node("extract_variables", extract_variables)
    workflow.add_node("check_conflict", check_conflict)
    workflow.add_node("check_availability", check_availability)
    workflow.add_node("output_schedule", output_schedule)
    workflow.add_node("ask_new_time", ask_for_new_time)

    workflow.set_entry_point("extract_variables")

    workflow.add_edge("extract_variables", "check_conflict")
    
    workflow.add_conditional_edges(
        "check_conflict",
        route_conflict_check,
        {
            "ask_new_time": "ask_new_time",
            "check_availability": "check_availability",
        }
    )

    workflow.add_conditional_edges(
        "check_availability",
        route_availability_check,
        {
            "output_schedule": "output_schedule",
            "ask_new_time": "ask_new_time",
        }
    )

    workflow.add_conditional_edges(
        "ask_new_time",
        route_after_ask_new_time,
        {
            "END": END,
            "extract_variables": "extract_variables",
        }
    )

    workflow.add_edge("output_schedule", END)

    return workflow.compile()

# ---------------------------------------------------------------------------
# 7) IMPROVED USER INTERFACE
# ---------------------------------------------------------------------------

def show_excel_tasks(file_path: str):
    """Show tasks from Excel"""
    try:
        df = read_all_tasks_from_excel(file_path)
        if df.empty:
            print("📅 No tasks in Excel yet.")
        else:
            print(f"\n📊 Tasks in Excel ({len(df)} tasks):")
            print("=" * 80)
            for idx, row in df.iterrows():
                print(f"{idx+1}. {row['Task Name']} | {row['Start Time']} | {row['Duration']} | {row['Date']}")
            print("=" * 80)
    except Exception as e:
        print(f"❌ Error displaying tasks: {e}")

def show_schedule(state):
    """Show schedule"""
    tasks = state.get("completed_tasks", []) + state.get("tasks", [])
    
    if not tasks:
        print("📅 No scheduled tasks.")
        return

    print("\n📅 Task Schedule:")
    print("=" * 70)
    for i, task in enumerate(tasks, start=1):
        name = task.get("name", "Not specified")
        time = task.get("start_time", "Not specified")
        duration = task.get("duration", "Not specified")
        status = task.get("status", "pending")
        print(f"{i}. {name} - {time} - {duration} - {status}")
    print("=" * 70)

def run_agent(graph):
    """Run agent"""
    print("=" * 50)
    print("🎯 Task Scheduling System (LangGraph)")
    print("=" * 50)
    
    excel_file = "schedule_tasks.xlsx"
    excel_path = initialize_excel_file(excel_file)
    
    show_excel_tasks(excel_path)
    
    while True:
        print("\n" + "="*50)
        user_input = input("📝 Enter scheduling request (or 'exit' to quit): ").strip()
        
        if user_input.lower() in ['exit', 'quit', 'q']:
            break
        
        if not user_input:
            print("❌ No input provided")
            continue

        initial_state = {
            "user_input": user_input,
            "tasks": [],
            "current_task_index": 0,
            "completed_tasks": [],
            "schedule_date": "",
            "conflict_detected": False,
            "is_available": True,
            "llm_response": "",
            "new_user_input": "",
            "extraction_failed": False,
            "excel_file_path": excel_path,
            "conflict_details": {},
            "max_retries": 3,
            "current_retry": 0,
            "waiting_for_user_input": False
        }
        
        try:
            config = {"recursion_limit": 10}
            final_state = None
            
            # First invocation
            state = graph.invoke(initial_state, config=config)
            final_state = state
            
            print("\n--- Result ---")
            print(state["llm_response"])
            
            # Handle conflicts with user input
            max_retries = 3
            retry_count = 0
            
            while state.get("waiting_for_user_input") and retry_count < max_retries:
                retry_count += 1
                
                new_input = input("⏰ Enter new time/date: ").strip()
                
                if not new_input:
                    print("⚠️ No new time entered, using original time")
                    new_input = user_input
                
                # Create new state with user input
                new_initial_state = {
                    "user_input": new_input,
                    "tasks": [],
                    "current_task_index": 0,
                    "completed_tasks": state.get("completed_tasks", []),
                    "schedule_date": "",
                    "conflict_detected": False,
                    "is_available": True,
                    "llm_response": "",
                    "new_user_input": "",
                    "extraction_failed": False,
                    "excel_file_path": excel_path,
                    "conflict_details": {},
                    "max_retries": max_retries,
                    "current_retry": retry_count,
                    "waiting_for_user_input": False
                }
                
                state = graph.invoke(new_initial_state, config=config)
                final_state = state
                
                print("\n--- Result ---")
                print(state["llm_response"])
                
                if not state.get("waiting_for_user_input"):
                    break

            if retry_count >= max_retries:
                print("\n🚫 Maximum attempts reached")
            
            print("\n✅ Process completed!")
            
            if final_state:
                show_schedule(final_state)
            show_excel_tasks(excel_path)
            
            print(f"\n💾 Excel file: {os.path.abspath(excel_path)}")
            
        except Exception as e:
            print(f"❌ Execution error: {e}")
            print("🔄 Restarting system...")

    print("\n👋 Thank you for using the Task Scheduling System!")

if __name__ == "__main__":
    try:
        graph = build_workflow()
        run_agent(graph)
    except Exception as e:
        print(f"❌ Runtime error: {e}")